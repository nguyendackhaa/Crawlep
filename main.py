import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import queue
import os
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from PIL import Image, ImageOps
import requests
import io
import time
from urllib.parse import urlparse
import re
from datetime import datetime
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ImageNamingProcessor:
    """Class xử lý đặt tên file ảnh theo logic từ JavaScript"""
    
    def __init__(self):
        self.domain = "https://example.com/product/"
        self.image_base = "https://cdn.example.com/images/"
    
    def standardize(self, text):
        """Chuẩn hóa chuỗi theo logic từ JavaScript"""
        if not text:
            return ""
        
        # Thay ký tự không hợp lệ bằng '-'
        standardized = re.sub(r'[\\/:*?"<>|,=\s]', '-', text)
        
        # Loại bỏ gạch ngang trùng
        standardized = re.sub(r'-+', '-', standardized)
        
        # Chỉ giữ lại a-z, A-Z, 0-9, -, _
        standardized = re.sub(r'[^a-zA-Z0-9\-_]', '', standardized)
        
        # Loại bỏ gạch đầu/cuối
        standardized = re.sub(r'^-+|-+$', '', standardized)
        
        return standardized
    
    def process_product_code(self, code):
        """Xử lý mã sản phẩm để tạo slug và image name"""
        if not code:
            return "", "", False
        
        # Kiểm tra có add-on kit không
        had_addon_kit = bool(re.search(r'add[\s\-]*on[\s\-]*kit', code, re.IGNORECASE))
        
        # Làm sạch chuỗi, loại bỏ ghi chú coating, addon...
        clean_code = code
        clean_code = re.sub(r'\(with special coating\)', '', clean_code, flags=re.IGNORECASE)
        clean_code = re.sub(r'\[with special coating\]', '', clean_code, flags=re.IGNORECASE)
        clean_code = re.sub(r'add[\s\-]*on[\s\-]*kit', '', clean_code, flags=re.IGNORECASE)
        clean_code = clean_code.strip()
        
        # Tạo slug (lowercase) và image name (uppercase)
        slug = self.standardize(clean_code.lower())
        if had_addon_kit:
            slug += "-adk"
        
        image_name = self.standardize(clean_code.upper())
        
        return slug, image_name, had_addon_kit
    
    def generate_filename(self, code):
        """Tạo tên file ảnh theo logic JavaScript"""
        slug, image_name, had_addon = self.process_product_code(code)
        return image_name + ".webp" if image_name else "unknown.webp"

class ImageCrawlerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Image Crawler - Cào Ảnh Tự Động")
        self.root.geometry("1000x700")
        self.root.configure(bg='#f0f0f0')
        
        # Khởi tạo queue cho đa luồng
        self.download_queue = queue.Queue()
        self.worker_threads = []
        self.max_workers = 5
        self.is_crawling = False
        
        # Dữ liệu Excel để mapping mã sản phẩm
        self.excel_data = None
        self.product_codes = []  # List of all entries from Excel - no duplicate filtering
        
        # Result tracking system cho Excel reporting
        self.results = []  # Detailed results for each entry
        self.start_time = None
        self.output_dir = None
        
        # Khởi tạo image naming processor
        self.naming_processor = ImageNamingProcessor()
        
        # Tạo giao diện
        self.create_widgets()
        
        # Khởi động worker threads
        self.start_worker_threads()
        
    def create_widgets(self):
        # Style cho giao diện
        style = ttk.Style()
        style.theme_use('clam')
        
        # Frame chính
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Cấu hình grid
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Tiêu đề
        title_label = ttk.Label(main_frame, text="🖼️ IMAGE CRAWLER - XỬ LÝ ẢNH SẢN PHẨM", 
                               font=('Arial', 16, 'bold'), foreground='#2c3e50')
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # Input section
        input_frame = ttk.LabelFrame(main_frame, text="Nhập Link Sản Phẩm", padding="10")
        input_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        input_frame.columnconfigure(1, weight=1)
        
        # Radio buttons cho input type
        self.input_type = tk.StringVar(value="excel")
        ttk.Radiobutton(input_frame, text="File Excel (Khuyến nghị)", variable=self.input_type, 
                       value="excel", command=self.toggle_input_type).grid(row=0, column=0, sticky=tk.W)
        ttk.Radiobutton(input_frame, text="Danh sách link", variable=self.input_type, 
                       value="list", command=self.toggle_input_type).grid(row=0, column=1, sticky=tk.W)
        
        # Text area cho links
        self.links_text = scrolledtext.ScrolledText(input_frame, height=6, width=60)
        self.links_text.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(10, 0))
        self.links_text.insert(tk.END, "Chọn file Excel để import links và mã sản phẩm...")
        self.links_text.config(state='disabled')
        
        # Browse button cho Excel
        self.browse_button = ttk.Button(input_frame, text="Chọn File Excel", 
                                       command=self.browse_excel)
        self.browse_button.grid(row=2, column=0, columnspan=3, pady=(10, 0))
        
        # Debug button
        self.debug_button = ttk.Button(input_frame, text="🔍 Debug Excel", 
                                      command=self.debug_excel_info, state='disabled')
        self.debug_button.grid(row=3, column=0, columnspan=3, pady=(5, 0))
        
        # Cấu hình crawler
        config_frame = ttk.LabelFrame(main_frame, text="Cấu Hình Xử Lý Ảnh", padding="10")
        config_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        config_frame.columnconfigure(1, weight=1)
        
        ttk.Label(config_frame, text="Số luồng xử lý:").grid(row=0, column=0, sticky=tk.W)
        self.thread_count = tk.StringVar(value="5")
        thread_spinbox = ttk.Spinbox(config_frame, from_=1, to=10, textvariable=self.thread_count, width=10)
        thread_spinbox.grid(row=0, column=1, sticky=tk.W, padx=(10, 0))
        
        ttk.Label(config_frame, text="Thư mục lưu:").grid(row=1, column=0, sticky=tk.W, pady=(10, 0))
        self.save_path = tk.StringVar(value="./downloaded_images")
        path_entry = ttk.Entry(config_frame, textvariable=self.save_path, width=50)
        path_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(10, 0))
        ttk.Button(config_frame, text="Chọn", command=self.browse_folder).grid(row=1, column=2, padx=(10, 0), pady=(10, 0))
        
        # Xử lý ảnh sản phẩm
        ttk.Label(config_frame, text="Xử lý ảnh:").grid(row=2, column=0, sticky=tk.W, pady=(10, 0))
        self.image_processing = tk.StringVar(value="product")
        process_frame = ttk.Frame(config_frame)
        process_frame.grid(row=2, column=1, sticky=tk.W, padx=(10, 0), pady=(10, 0))
        ttk.Radiobutton(process_frame, text="Ảnh sản phẩm (có nền trắng)", variable=self.image_processing, 
                       value="product").pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(process_frame, text="Ảnh thường", variable=self.image_processing, 
                       value="normal").pack(side=tk.LEFT)
        
        # Crawl mode
        ttk.Label(config_frame, text="Chế độ crawl:").grid(row=3, column=0, sticky=tk.W, pady=(10, 0))
        self.crawl_mode = tk.StringVar(value="direct")
        crawl_frame = ttk.Frame(config_frame)
        crawl_frame.grid(row=3, column=1, sticky=tk.W, padx=(10, 0), pady=(10, 0))
        ttk.Radiobutton(crawl_frame, text="Link ảnh trực tiếp", variable=self.crawl_mode, 
                       value="direct").pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(crawl_frame, text="Crawl từ trang web", variable=self.crawl_mode, 
                       value="webpage").pack(side=tk.LEFT)
        
        # Control buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=3, column=0, columnspan=3, pady=20)
        
        self.start_button = ttk.Button(button_frame, text="🚀 Bắt Đầu Crawl", 
                                       command=self.start_crawling, style='Accent.TButton')
        self.start_button.pack(side=tk.LEFT, padx=(0, 10))
        
        self.stop_button = ttk.Button(button_frame, text="⏹️ Dừng", 
                                     command=self.stop_crawling, state='disabled')
        self.stop_button.pack(side=tk.LEFT, padx=(0, 10))
        
                # Test button
        self.test_button = ttk.Button(button_frame, text="🧪 Test với Links Mẫu", 
                                      command=self.load_test_links)
        self.test_button.pack(side=tk.LEFT, padx=(0, 10))
        
        # Test naming button
        self.test_naming_button = ttk.Button(button_frame, text="📝 Test Đặt Tên", 
                                            command=self.test_naming_logic)
        self.test_naming_button.pack(side=tk.LEFT)
        
        # Progress section
        progress_frame = ttk.LabelFrame(main_frame, text="Tiến Trình", padding="10")
        progress_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        progress_frame.columnconfigure(0, weight=1)
        progress_frame.rowconfigure(2, weight=1)
        
        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Status label
        self.status_label = ttk.Label(progress_frame, text="Sẵn sàng để bắt đầu...")
        self.status_label.grid(row=1, column=0, sticky=tk.W, pady=(0, 10))
        
        # Log area
        self.log_text = scrolledtext.ScrolledText(progress_frame, height=10, width=80)
        self.log_text.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Stats
        stats_frame = ttk.Frame(progress_frame)
        stats_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(10, 0))
        
        self.total_links_label = ttk.Label(stats_frame, text="Tổng link: 0")
        self.total_links_label.pack(side=tk.LEFT, padx=(0, 20))
        
        self.processed_label = ttk.Label(stats_frame, text="Đã xử lý: 0")
        self.processed_label.pack(side=tk.LEFT, padx=(0, 20))
        
        self.success_label = ttk.Label(stats_frame, text="Thành công: 0")
        self.success_label.pack(side=tk.LEFT, padx=(0, 20))
        
        self.failed_label = ttk.Label(stats_frame, text="Thất bại: 0")
        self.failed_label.pack(side=tk.LEFT)
        
        # Cấu hình grid weights
        main_frame.rowconfigure(4, weight=1)
        
    def toggle_input_type(self):
        if self.input_type.get() == "list":
            self.links_text.config(state='normal')
            self.browse_button.config(state='disabled')
            self.links_text.delete(1.0, tk.END)
            self.links_text.insert(tk.END, "Nhập các link sản phẩm, mỗi link một dòng...")
        else:
            self.links_text.config(state='disabled')
            self.browse_button.config(state='normal')
            self.links_text.delete(1.0, tk.END)
            self.links_text.insert(tk.END, "Chọn file Excel để import links và mã sản phẩm...")
    
    def load_test_links(self):
        """Load test links từ file test_links.txt"""
        try:
            if os.path.exists("test_links.txt"):
                with open("test_links.txt", "r", encoding="utf-8") as f:
                    test_links = f.read()
                
                self.input_type.set("list")
                self.toggle_input_type()
                self.links_text.delete(1.0, tk.END)
                self.links_text.insert(tk.END, test_links)
                self.log_message("Đã load test links thành công!")
            else:
                messagebox.showwarning("Cảnh báo", "Không tìm thấy file test_links.txt")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể load test links: {str(e)}")
    
    def test_naming_logic(self):
        """Test logic đặt tên file ảnh theo JavaScript"""
        test_codes = [
            "FR-1H-220V",
            "TC-NT-20R (with special coating)",
            "ETC-48 Add-on Kit",
            "FR-2H-380V [with special coating]",
            "TC-NT-30R ADD ON KIT",
            "ETC-60",
            "FR-3H-440V"
        ]
        
        self.log_message("🧪 TEST LOGIC ĐẶT TÊN FILE ẢNH")
        self.log_message("=" * 50)
        
        for code in test_codes:
            filename = self.naming_processor.generate_filename(code)
            slug, image_name, had_addon = self.naming_processor.process_product_code(code)
            
            self.log_message(f"📋 Mã gốc: {code}")
            self.log_message(f"🧹 Mã sạch: {slug}")
            self.log_message(f"🖼️ Tên ảnh: {image_name}")
            self.log_message(f"📦 Add-on kit: {had_addon}")
            self.log_message(f"📄 Filename: {filename}")
            self.log_message("-" * 30)
        
        self.log_message("✅ Hoàn thành test logic đặt tên!")
        messagebox.showinfo("Test Hoàn Thành", "Đã test logic đặt tên file ảnh!\nXem log để biết chi tiết.")
    
    def browse_excel(self):
        filename = filedialog.askopenfilename(
            title="Chọn File Excel",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            try:
                # Đọc file Excel với nhiều cách khác nhau để đảm bảo nhận đầy đủ dữ liệu
                try:
                    # Thử đọc bình thường trước
                    df = pd.read_excel(filename)
                    self.log_message(f"Đọc bình thường: {len(df)} dòng")
                except:
                    df = pd.read_excel(filename, header=None)
                    self.log_message(f"Đọc không header: {len(df)} dòng")
                
                # Thử đọc với header=None để so sánh
                try:
                    df_no_header = pd.read_excel(filename, header=None)
                    if len(df_no_header) > len(df):
                        df = df_no_header
                        self.log_message(f"Chuyển sang đọc không header: {len(df)} dòng")
                except:
                    pass
                
                # Lưu dữ liệu để debug
                self.last_excel_data = df
                
                # Hiển thị thông tin file
                self.log_message(f"📊 Thông tin file Excel: {filename}")
                self.log_message(f"📋 Tổng số cột: {len(df.columns)}")
                self.log_message(f"📝 Tổng số dòng: {len(df)}")
                
                # Kiểm tra cấu trúc file
                if len(df.columns) < 2:
                    messagebox.showerror("Lỗi", "File Excel phải có ít nhất 2 cột:\nCột A: Mã sản phẩm\nCột B: Link ảnh")
                    return
                
                # Lấy cột A (mã sản phẩm) và cột B (link)
                product_codes = df.iloc[:, 0].tolist()
                links = df.iloc[:, 1].tolist()
                
                # Hiển thị dữ liệu gốc
                self.log_message(f"📋 Dữ liệu gốc từ Excel:")
                for i, (code, link) in enumerate(zip(product_codes, links)):
                    self.log_message(f"   Dòng {i+1}: Mã='{code}' | Link='{link}'")
                
                # Tạo list entries (LOGIC ĐƠN GIẢN - KHÔNG PHÂN BIỆT DUPLICATE)
                self.product_codes = []
                valid_count = 0
                skipped_count = 0
                
                self.log_message(f"🔧 BẮT ĐẦU XỬ LÝ {len(product_codes)} DÒNG DỮ LIỆU")
                
                for i, (code, link) in enumerate(zip(product_codes, links)):
                    # Kiểm tra và xử lý dữ liệu
                    code_str = str(code).strip() if pd.notna(code) else ""
                    link_str = str(link).strip() if pd.notna(link) else ""
                    
                    # Debug chi tiết từng dòng
                    self.log_message(f"🔍 Dòng {i+1}: Mã='{code_str}' | Link='{link_str}'")
                    
                    # LOGIC ĐƠN GIẢN: Chỉ bỏ qua dòng hoàn toàn trống
                    if not code_str and not link_str:
                        self.log_message(f"⚠️ Bỏ qua dòng {i+1}: Dòng trống hoàn toàn")
                        skipped_count += 1
                        continue
                    
                    # Nếu có link nhưng không có mã, tự tạo mã
                    if link_str and not code_str:
                        code_str = f"PRODUCT_{i+1:03d}"
                        self.log_message(f"⚠️ Dòng {i+1}: Tự tạo mã '{code_str}' cho link")
                    
                    # Nếu có mã nhưng không có link, tự tạo link
                    if code_str and not link_str:
                        link_str = f"https://example.com/product/{code_str}"
                        self.log_message(f"⚠️ Dòng {i+1}: Tự tạo link cho mã '{code_str}'")
                    
                    # THÊM TẤT CẢ ENTRIES VÀO LIST (KHÔNG PHÂN BIỆT DUPLICATE)
                    if link_str and code_str:
                        # Tạo entry object đơn giản
                        entry = {
                            'code': code_str,
                            'link': link_str,
                            'row': i + 1
                        }
                        
                        # Thêm vào list - không kiểm tra duplicate
                        self.product_codes.append(entry)
                        valid_count += 1
                        self.log_message(f"✅ Dòng {i+1}: Mã='{code_str}' | Link='{link_str}' | Entry #{len(self.product_codes)}")
                    else:
                        self.log_message(f"⚠️ Dòng {i+1}: Thiếu thông tin")
                        skipped_count += 1
                
                # Hiển thị kết quả
                total_entries = len(self.product_codes)
                
                self.log_message(f"📊 Kết quả xử lý Excel:")
                self.log_message(f"   ✅ Tổng entries: {total_entries}")
                self.log_message(f"   ⚠️ Bỏ qua: {skipped_count} dòng")
                self.log_message(f"   📋 Tổng cộng: {len(df)} dòng")
                
                # Hiển thị trong text area (LOGIC ĐƠN GIẢN)
                self.links_text.config(state='normal')
                self.links_text.delete(1.0, tk.END)
                
                # Hiển thị tất cả entries
                for idx, entry in enumerate(self.product_codes):
                    self.links_text.insert(tk.END, f"{entry['code']}\t{entry['link']}\n")
                
                self.links_text.config(state='disabled')
                
                self.log_message(f"🎯 Đã import {total_entries} entries từ file Excel")
                self.log_message(f"📋 Cột A: Mã sản phẩm, Cột B: Link ảnh")
                
                # Kích hoạt nút debug
                self.debug_button.config(state='normal')
                
                # Hiển thị thông báo đơn giản
                if total_entries > 0:
                    messagebox.showinfo("Thành công", 
                        f"Đã import thành công {total_entries} entries!\n\n"
                        f"📊 Thống kê:\n"
                        f"- Tổng dòng Excel: {len(df)}\n"
                        f"- Entries hợp lệ: {total_entries}\n"
                        f"- Dòng bỏ qua: {skipped_count}\n\n"
                        f"App sẽ xử lý tất cả {total_entries} entries!")
                else:
                    messagebox.showwarning("Cảnh báo", 
                        f"Không có dữ liệu hợp lệ nào được tìm thấy!\n\n"
                        f"Tổng dòng Excel: {len(df)}\n"
                        f"Dòng bỏ qua: {skipped_count}\n\n"
                        f"Vui lòng kiểm tra cấu trúc file Excel.")
                
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể đọc file Excel: {str(e)}")
                self.log_message(f"❌ Lỗi khi đọc Excel: {str(e)}")
    
    def debug_excel_info(self):
        """Hiển thị thông tin chi tiết về file Excel đã import"""
        if not hasattr(self, 'last_excel_data') or self.last_excel_data is None:
            messagebox.showinfo("Thông tin", "Chưa có file Excel nào được import!")
            return
        
        try:
            df = self.last_excel_data
            
            # Tạo thông tin debug
            debug_info = f"🔍 THÔNG TIN DEBUG FILE EXCEL\n"
            debug_info += f"{'='*50}\n\n"
            debug_info += f"📊 Thông tin chung:\n"
            debug_info += f"   - Tổng số cột: {len(df.columns)}\n"
            debug_info += f"   - Tổng số dòng: {len(df)}\n"
            debug_info += f"   - Tên cột: {list(df.columns)}\n\n"
            
            debug_info += f"📋 Dữ liệu chi tiết:\n"
            for i in range(len(df)):
                col_a = df.iloc[i, 0] if len(df.columns) > 0 else "N/A"
                col_b = df.iloc[i, 1] if len(df.columns) > 1 else "N/A"
                
                col_a_str = str(col_a) if pd.notna(col_a) else "TRỐNG"
                col_b_str = str(col_b) if pd.notna(col_b) else "TRỐNG"
                
                debug_info += f"   Dòng {i+1:2d}: | {col_a_str:<20} | {col_b_str}\n"
            
            debug_info += f"\n🎯 Kết quả xử lý:\n"
            total_entries = len(self.product_codes)
            
            debug_info += f"   - Tổng entries: {total_entries}\n"
            debug_info += f"   - Mapping chi tiết:\n"
            
            for idx, entry in enumerate(self.product_codes):
                debug_info += f"     {idx+1:2d}. {entry['code']} (row {entry['row']}) -> {entry['link']}\n"
            
            # Hiển thị trong dialog
            self.show_debug_dialog(debug_info)
            
        except Exception as e:
            messagebox.showerror("Lỗi Debug", f"Không thể hiển thị thông tin debug: {str(e)}")
    
    def show_debug_dialog(self, debug_info):
        """Hiển thị dialog debug với thông tin chi tiết"""
        debug_window = tk.Toplevel(self.root)
        debug_window.title("🔍 Debug Excel Info")
        debug_window.geometry("800x600")
        debug_window.configure(bg='#f0f0f0')
        
        # Frame chính
        main_frame = ttk.Frame(debug_window, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Tiêu đề
        title_label = ttk.Label(main_frame, text="🔍 THÔNG TIN DEBUG FILE EXCEL", 
                               font=('Arial', 14, 'bold'))
        title_label.pack(pady=(0, 10))
        
        # Text area cho debug info
        debug_text = scrolledtext.ScrolledText(main_frame, height=25, width=90, font=('Consolas', 10))
        debug_text.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        debug_text.insert(tk.END, debug_info)
        debug_text.config(state='disabled')
        
        # Nút đóng
        close_button = ttk.Button(main_frame, text="Đóng", command=debug_window.destroy)
        close_button.pack()
        
        # Focus vào window
        debug_window.focus_set()
        debug_window.grab_set()  # Modal window
    
    def browse_folder(self):
        folder = filedialog.askdirectory(title="Chọn thư mục lưu ảnh")
        if folder:
            self.save_path.set(folder)
    
    def start_worker_threads(self):
        for i in range(self.max_workers):
            worker = threading.Thread(target=self.worker_function, daemon=True)
            worker.start()
            self.worker_threads.append(worker)
    
    def worker_function(self):
        while True:
            try:
                task = self.download_queue.get(timeout=1)
                if task is None:
                    break
                
                # Handle both old 3-param and new 4-param format for backward compatibility
                if len(task) == 4:
                    link, save_dir, product_code, row_number = task
                else:
                    link, save_dir, product_code = task
                    row_number = None
                self.process_single_link(link, save_dir, product_code, row_number)
                self.download_queue.task_done()
                
            except queue.Empty:
                continue
            except Exception as e:
                self.log_message(f"Lỗi worker thread: {str(e)}")
    
    def start_crawling(self):
        if self.is_crawling:
            return
        
        # Lấy entries hoặc links
        if self.input_type.get() == "excel":
            if not self.product_codes:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn file Excel trước!")
                return
            # Sử dụng entries thay vì links để xử lý tất cả
            entries = self.product_codes
        else:
            links_text = self.links_text.get(1.0, tk.END).strip()
            if not links_text or links_text == "Nhập các link sản phẩm, mỗi link một dòng...":
                messagebox.showwarning("Cảnh báo", "Vui lòng nhập ít nhất một link!")
                return
            links = [link.strip() for link in links_text.split('\n') if link.strip() and not link.startswith('#')]
            # Convert links to entries format
            entries = [{'code': f'manual_{i+1}', 'link': link, 'row': i+1} for i, link in enumerate(links)]
        
        if not entries:
            messagebox.showwarning("Cảnh báo", "Không có entry hợp lệ nào!")
            return
        
        # Tạo thư mục lưu
        save_dir = self.save_path.get()
        os.makedirs(save_dir, exist_ok=True)
        
        # Cập nhật UI
        self.is_crawling = True
        self.start_button.config(state='disabled')
        self.stop_button.config(state='normal')
        self.progress_var.set(0)
        
        # Reset stats và khởi tạo tracking
        self.total_links = len(entries)
        self.processed_count = 0
        self.success_count = 0
        self.failed_count = 0
        self.results = []  # Reset results tracking
        self.start_time = time.time()  # Set start time for reporting
        self.output_dir = save_dir  # Store output directory
        
        self.update_stats()
        self.log_message(f"Bắt đầu crawl {len(entries)} entries...")
        
        # Bắt đầu crawl trong thread riêng - truyền entries thay vì links
        crawl_thread = threading.Thread(target=self.crawl_entries, args=(entries, save_dir))
        crawl_thread.start()
    
    def crawl_entries(self, entries, save_dir):
        try:
            if self.crawl_mode.get() == "direct":
                # Chế độ link ảnh trực tiếp
                self.log_message("Chế độ: Link ảnh trực tiếp")
                for i, entry in enumerate(entries):
                    if not self.is_crawling:
                        break
                    
                    try:
                        link = entry['link']
                        product_code = entry['code']
                        row = entry['row']
                        
                        self.log_message(f"Đang xử lý entry {i+1}/{len(entries)}: {product_code} -> {link} (row {row})")
                        
                        # Thêm vào queue download - XỬ LÝ TỪNG ENTRY
                        self.download_queue.put((link, save_dir, product_code, row))
                        self.log_message(f"Entry được thêm vào queue: {product_code} -> {link}")
                        
                        # Cập nhật progress
                        progress = ((i + 1) / len(entries)) * 100
                        self.root.after(0, lambda p=progress: self.progress_var.set(p))
                        
                    except Exception as e:
                        self.log_message(f"Lỗi khi xử lý entry {entry}: {str(e)}")
                        self.failed_count += 1
                    
                    self.processed_count += 1
                    self.update_stats()
            else:
                # Chế độ crawl từ trang web
                self.log_message("Chế độ: Crawl từ trang web")
                service = Service(ChromeDriverManager().install())
                options = webdriver.ChromeOptions()
                options.add_argument('--headless')
                options.add_argument('--no-sandbox')
                options.add_argument('--disable-dev-shm-usage')
                
                driver = webdriver.Chrome(service=service, options=options)
                
                for i, entry in enumerate(entries):
                    if not self.is_crawling:
                        break
                    
                    try:
                        link = entry['link']
                        product_code = entry['code']
                        row = entry['row']
                        
                        self.log_message(f"Đang xử lý entry {i+1}/{len(entries)}: {product_code} -> {link} (row {row})")
                        
                        # Crawl ảnh từ link
                        images = self.crawl_images_from_link(driver, link)
                        
                        if images:
                            # Thêm vào queue download
                            for img_url in images:
                                self.download_queue.put((img_url, save_dir, product_code, row))
                            
                            self.log_message(f"Tìm thấy {len(images)} ảnh từ entry: {product_code}")
                        else:
                            self.log_message(f"Không tìm thấy ảnh nào từ entry: {product_code}")
                        
                        # Cập nhật progress
                        progress = ((i + 1) / len(entries)) * 100
                        self.root.after(0, lambda p=progress: self.progress_var.set(p))
                        
                    except Exception as e:
                        self.log_message(f"Lỗi khi xử lý entry {entry}: {str(e)}")
                        self.failed_count += 1
                    
                    self.processed_count += 1
                    self.update_stats()
                
                driver.quit()
            
            # Chờ tất cả download hoàn thành
            self.download_queue.join()
            
            self.root.after(0, self.crawling_finished)
            
        except Exception as e:
            self.log_message(f"Lỗi trong quá trình crawl: {str(e)}")
            self.root.after(0, self.crawling_finished)
    
    def crawl_images_from_link(self, driver, link):
        try:
            driver.get(link)
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "img"))
            )
            
            # Tìm tất cả ảnh
            images = driver.find_elements(By.TAG_NAME, "img")
            image_urls = []
            
            for img in images:
                src = img.get_attribute('src')
                if src and self.is_valid_image_url(src):
                    image_urls.append(src)
            
            return image_urls
            
        except Exception as e:
            self.log_message(f"Lỗi khi crawl link {link}: {str(e)}")
            return []
    
    def is_valid_image_url(self, url):
        if not url:
            return False
        
        # Kiểm tra xem có phải là URL không
        if not url.startswith(('http://', 'https://')):
            return False
        
        # Kiểm tra extension
        valid_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.JPG', '.JPEG', '.PNG', '.GIF', '.BMP', '.WEBP']
        parsed_url = urlparse(url)
        path = parsed_url.path.lower()
        
        # Kiểm tra extension trong path
        has_valid_extension = any(path.endswith(ext.lower()) for ext in valid_extensions)
        
        # Kiểm tra query parameters có chứa extension
        query_has_extension = any(ext.lower() in parsed_url.query.lower() for ext in valid_extensions)
        
        # Kiểm tra fragment có chứa extension
        fragment_has_extension = any(ext.lower() in parsed_url.fragment.lower() for ext in valid_extensions)
        
        # Nếu có extension ở bất kỳ đâu, coi như hợp lệ
        if has_valid_extension or query_has_extension or fragment_has_extension:
            return True
        
        # Kiểm tra một số pattern đặc biệt
        special_patterns = [
            'cdn', 'images', 'img', 'photo', 'picture', 'upload', 'media',
            'static', 'assets', 'content', 'files', 'storage'
        ]
        
        url_lower = url.lower()
        has_special_pattern = any(pattern in url_lower for pattern in special_patterns)
        
        # Nếu có pattern đặc biệt, coi như hợp lệ
        if has_special_pattern:
            return True
        
        return False
    
    def process_single_link(self, img_url, save_dir, product_code, row_number=None):
        # Initialize result tracking
        start_time = time.time()
        result_entry = {
            'product_code': product_code,
            'link': img_url, 
            'row': row_number,
            'status': 'failed',
            'filename': None,
            'file_size': None,
            'error_reason': None,
            'download_time': None,
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        try:
            # Kiểm tra xem có phải link ảnh trực tiếp không
            if self.is_valid_image_url(img_url):
                # Download ảnh trực tiếp
                self.log_message(f"🖼️ Download ảnh trực tiếp: {img_url}")
                
                try:
                    # Enhanced headers để bypass 403 Forbidden
                    headers = {
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
                        'Referer': 'https://www.fotekexpress.com/',
                        'Accept': 'image/webp,image/apng,image/*,*/*;q=0.8',
                        'Accept-Language': 'en-US,en;q=0.9',
                        'Accept-Encoding': 'gzip, deflate, br',
                        'Connection': 'keep-alive',
                        'Upgrade-Insecure-Requests': '1',
                    }
                    response = requests.get(img_url, headers=headers, timeout=30)
                    response.raise_for_status()
                    
                    # Xử lý ảnh
                    img = Image.open(io.BytesIO(response.content))
                    
                    if self.image_processing.get() == "product":
                        # Xử lý ảnh sản phẩm: chèn nền trắng
                        img = self.process_product_image(img)
                    
                    # Convert RGBA sang RGB nếu cần
                    if img.mode in ('RGBA', 'LA', 'P'):
                        img = img.convert('RGB')
                    
                    # Tạo tên file theo mã sản phẩm
                    filename = self.generate_filename(product_code)
                    filepath = os.path.join(save_dir, filename)
                    
                    # Lưu dưới dạng WebP
                    img.save(filepath, 'WEBP', quality=85, optimize=True)
                    
                    # Lấy file size
                    file_size = os.path.getsize(filepath)
                    
                    # Update result entry for success
                    result_entry.update({
                        'status': 'success',
                        'filename': filename,
                        'file_size': file_size,
                        'download_time': time.time() - start_time
                    })
                    
                    self.success_count += 1
                    self.log_message(f"✅ Đã lưu ảnh: {filename} (Mã: {product_code}) - {file_size/1024:.1f}KB")
                    
                except requests.exceptions.Timeout:
                    result_entry['error_reason'] = "Timeout - Link không phản hồi trong 30s"
                    self.failed_count += 1
                    self.log_message(f"❌ Timeout khi download: {img_url}")
                    
                except requests.exceptions.HTTPError as e:
                    result_entry['error_reason'] = f"HTTP Error {e.response.status_code}: {e.response.reason}"
                    self.failed_count += 1
                    self.log_message(f"❌ HTTP Error {e.response.status_code}: {img_url}")
                    
                except requests.exceptions.RequestException as e:
                    result_entry['error_reason'] = f"Network Error: {str(e)}"
                    self.failed_count += 1
                    self.log_message(f"❌ Network Error: {img_url}")
                    
                except Exception as e:
                    result_entry['error_reason'] = f"Image Processing Error: {str(e)}"
                    self.failed_count += 1
                    self.log_message(f"❌ Image Error: {img_url} - {str(e)}")
                
            else:
                # Link không phải ảnh trực tiếp - thử crawl từ trang web
                self.log_message(f"🌐 Thử crawl từ trang web: {img_url}")
                try:
                    # Sử dụng Selenium để crawl
                    service = Service(ChromeDriverManager().install())
                    options = webdriver.ChromeOptions()
                    options.add_argument('--headless')
                    options.add_argument('--no-sandbox')
                    options.add_argument('--disable-dev-shm-usage')
                    
                    driver = webdriver.Chrome(service=service, options=options)
                    
                    # Crawl ảnh từ trang web
                    images = self.crawl_images_from_link(driver, img_url)
                    
                    if images:
                        # Lưu ảnh đầu tiên tìm được
                        img_url_direct = images[0]
                        self.log_message(f"🖼️ Tìm thấy ảnh: {img_url_direct}")
                        
                        # Download ảnh với enhanced headers
                        headers = {
                            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
                            'Referer': 'https://www.fotekexpress.com/',
                            'Accept': 'image/webp,image/apng,image/*,*/*;q=0.8',
                            'Accept-Language': 'en-US,en;q=0.9',
                            'Accept-Encoding': 'gzip, deflate, br',
                            'Connection': 'keep-alive',
                            'Upgrade-Insecure-Requests': '1',
                        }
                        response = requests.get(img_url_direct, headers=headers, timeout=30)
                        response.raise_for_status()
                        
                        # Xử lý ảnh
                        img = Image.open(io.BytesIO(response.content))
                        
                        if self.image_processing.get() == "product":
                            # Xử lý ảnh sản phẩm: chèn nền trắng
                            img = self.process_product_image(img)
                        
                        # Convert RGBA sang RGB nếu cần
                        if img.mode in ('RGBA', 'LA', 'P'):
                            img = img.convert('RGB')
                        
                        # Tạo tên file theo mã sản phẩm
                        filename = self.generate_filename(product_code)
                        filepath = os.path.join(save_dir, filename)
                        
                        # Lưu dưới dạng WebP
                        img.save(filepath, 'WEBP', quality=85, optimize=True)
                        
                        # Lấy file size
                        file_size = os.path.getsize(filepath)
                        
                        # Update result entry for success
                        result_entry.update({
                            'status': 'success',
                            'filename': filename,
                            'file_size': file_size,
                            'download_time': time.time() - start_time
                        })
                        
                        self.success_count += 1
                        self.log_message(f"✅ Đã lưu ảnh từ trang web: {filename} (Mã: {product_code}) - {file_size/1024:.1f}KB")
                    else:
                        result_entry['error_reason'] = "Không tìm thấy ảnh nào trên trang web"
                        self.log_message(f"⚠️ Không tìm thấy ảnh nào từ trang web: {img_url}")
                        self.failed_count += 1
                    
                    driver.quit()
                    
                except Exception as e:
                    result_entry['error_reason'] = f"Web Crawl Error: {str(e)}"
                    self.log_message(f"❌ Lỗi khi crawl từ trang web {img_url}: {str(e)}")
                    self.failed_count += 1
            
        except Exception as e:
            result_entry['error_reason'] = f"General Error: {str(e)}"
            self.failed_count += 1
            self.log_message(f"❌ Lỗi khi xử lý link {img_url}: {str(e)}")
        
        finally:
            # Ensure download time is set
            if result_entry['download_time'] is None:
                result_entry['download_time'] = time.time() - start_time
            
            # Add result to tracking list
            self.results.append(result_entry)
            self.update_stats()
    
    def process_product_image(self, img):
        """Xử lý ảnh sản phẩm: chèn nền trắng và giữ nguyên kích thước"""
        try:
            # Lấy kích thước gốc
            original_width, original_height = img.size
            
            # Tạo ảnh nền trắng với kích thước gốc
            white_bg = Image.new('RGB', (original_width, original_height), (255, 255, 255))
            
            # Convert ảnh gốc sang RGBA nếu cần
            if img.mode != 'RGBA':
                img = img.convert('RGBA')
            
            # Paste ảnh gốc lên nền trắng
            white_bg.paste(img, (0, 0), img)
            
            return white_bg
            
        except Exception as e:
            self.log_message(f"Lỗi khi xử lý ảnh sản phẩm: {str(e)}")
            return img  # Trả về ảnh gốc nếu có lỗi
    
    def generate_filename(self, product_code):
        """Tạo tên file theo logic JavaScript"""
        return self.naming_processor.generate_filename(str(product_code))
    
    def generate_excel_report(self, output_dir):
        """Tạo Excel report với color coding cho results"""
        try:
            self.log_message("📊 Đang tạo Excel report...")
            
            # Tạo workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # === SHEET 1: DETAILED RESULTS ===
            details_ws = wb.create_sheet("Chi Tiết Kết Quả")
            
            # Headers
            headers = [
                'STT', 'Mã Sản Phẩm', 'Link', 'Trạng Thái', 'Tên File', 
                'Kích Thước (KB)', 'Lý Do Lỗi', 'Thời Gian DL (s)', 'Row Excel', 'Timestamp'
            ]
            
            # Style definitions
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            failed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # Add headers
            for col, header in enumerate(headers, 1):
                cell = details_ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_alignment
            
            # Add data rows
            for idx, result in enumerate(self.results, 2):
                # STT
                details_ws.cell(row=idx, column=1, value=idx-1)
                
                # Mã Sản Phẩm
                details_ws.cell(row=idx, column=2, value=result['product_code'])
                
                # Link
                details_ws.cell(row=idx, column=3, value=result['link'])
                
                # Trạng Thái
                status_cell = details_ws.cell(row=idx, column=4, value=result['status'].upper())
                status_cell.font = Font(bold=True)
                
                # Tên File
                details_ws.cell(row=idx, column=5, value=result['filename'] or 'N/A')
                
                # Kích Thước
                if result['file_size']:
                    size_kb = round(result['file_size'] / 1024, 1)
                    details_ws.cell(row=idx, column=6, value=size_kb)
                else:
                    details_ws.cell(row=idx, column=6, value='N/A')
                
                # Lý Do Lỗi
                details_ws.cell(row=idx, column=7, value=result['error_reason'] or 'N/A')
                
                # Thời Gian Download
                if result['download_time']:
                    dl_time = round(result['download_time'], 2)
                    details_ws.cell(row=idx, column=8, value=dl_time)
                else:
                    details_ws.cell(row=idx, column=8, value='N/A')
                
                # Row Excel
                details_ws.cell(row=idx, column=9, value=result['row'] or 'N/A')
                
                # Timestamp
                details_ws.cell(row=idx, column=10, value=result['timestamp'])
                
                # Apply row coloring and borders
                for col in range(1, 11):
                    cell = details_ws.cell(row=idx, column=col)
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')
                    
                    if result['status'] == 'success':
                        cell.fill = success_fill
                    else:
                        cell.fill = failed_fill
            
            # Auto-adjust column widths
            for col in range(1, 11):
                column_letter = get_column_letter(col)
                max_length = 0
                for row in details_ws[f'{column_letter}1:{column_letter}{len(self.results)+1}']:
                    for cell in row:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = min(max_length + 2, 50)
                details_ws.column_dimensions[column_letter].width = adjusted_width
            
            # === SHEET 2: SUMMARY STATISTICS ===
            summary_ws = wb.create_sheet("Tổng Kết")
            
            # Calculate statistics
            total_entries = len(self.results)
            success_count = len([r for r in self.results if r['status'] == 'success'])
            failed_count = total_entries - success_count
            success_rate = (success_count / total_entries * 100) if total_entries > 0 else 0
            
            # Processing time
            if self.start_time:
                total_time = time.time() - self.start_time
                avg_time_per_entry = total_time / total_entries if total_entries > 0 else 0
            else:
                total_time = 0
                avg_time_per_entry = 0
            
            # Error breakdown
            error_breakdown = {}
            for result in self.results:
                if result['status'] == 'failed' and result['error_reason']:
                    error_type = result['error_reason'].split(':')[0]
                    error_breakdown[error_type] = error_breakdown.get(error_type, 0) + 1
            
            # Summary content
            summary_data = [
                ['📊 BÁO CÁO TỔNG KẾT CRAWLER', ''],
                ['', ''],
                ['Thống Kê Chung', ''],
                ['Tổng số entries', total_entries],
                ['Thành công', success_count],
                ['Thất bại', failed_count],
                ['Tỷ lệ thành công', f'{success_rate:.1f}%'],
                ['', ''],
                ['Thời Gian Xử Lý', ''],
                ['Tổng thời gian', f'{total_time:.1f}s'],
                ['Trung bình/entry', f'{avg_time_per_entry:.2f}s'],
                ['', ''],
                ['Phân Tích Lỗi', ''],
            ]
            
            # Add error breakdown
            for error_type, count in error_breakdown.items():
                summary_data.append([error_type, count])
            
            # Add summary data
            for row_idx, (label, value) in enumerate(summary_data, 1):
                summary_ws.cell(row=row_idx, column=1, value=label)
                summary_ws.cell(row=row_idx, column=2, value=value)
                
                # Style headers
                if label in ['📊 BÁO CÁO TỔNG KẾT CRAWLER', 'Thống Kê Chung', 'Thời Gian Xử Lý', 'Phân Tích Lỗi']:
                    summary_ws.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
                    summary_ws.cell(row=row_idx, column=1).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            
            # Auto-adjust summary columns
            summary_ws.column_dimensions['A'].width = 25
            summary_ws.column_dimensions['B'].width = 15
            
            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_filename = f"crawler_report_{timestamp}.xlsx"
            report_path = os.path.join(output_dir, report_filename)
            
            wb.save(report_path)
            
            self.log_message(f"✅ Đã tạo Excel report: {report_filename}")
            return report_path
            
        except Exception as e:
            self.log_message(f"❌ Lỗi khi tạo Excel report: {str(e)}")
            return None
    
    def create_output_package(self, base_save_dir):
        """Tạo organized output package với folder structure và files"""
        try:
            self.log_message("📁 Đang tạo output package...")
            
            # Tạo timestamp cho folder
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            package_name = f"crawler_output_{timestamp}"
            package_dir = os.path.join(base_save_dir, package_name)
            
            # Tạo folder structure
            images_dir = os.path.join(package_dir, "images")
            os.makedirs(images_dir, exist_ok=True)
            
            self.log_message(f"📁 Tạo folder: {package_dir}")
            
            # Copy successful images to images folder
            copied_count = 0
            for result in self.results:
                if result['status'] == 'success' and result['filename']:
                    source_path = os.path.join(base_save_dir, result['filename'])
                    dest_path = os.path.join(images_dir, result['filename'])
                    
                    try:
                        if os.path.exists(source_path):
                            # Copy file thay vì move để giữ nguyên file gốc
                            import shutil
                            shutil.copy2(source_path, dest_path)
                            copied_count += 1
                    except Exception as e:
                        self.log_message(f"⚠️ Không thể copy {result['filename']}: {str(e)}")
            
            self.log_message(f"📁 Đã copy {copied_count} ảnh vào folder images/")
            
            # Generate Excel report
            excel_path = self.generate_excel_report(package_dir)
            
            # Generate text summary
            summary_path = self.generate_text_summary(package_dir)
            
            # Generate package info
            package_info = {
                'package_dir': package_dir,
                'package_name': package_name,
                'images_dir': images_dir,
                'excel_path': excel_path,
                'summary_path': summary_path,
                'total_files': copied_count + (2 if excel_path and summary_path else 1 if excel_path or summary_path else 0),
                'images_count': copied_count
            }
            
            self.log_message(f"✅ Tạo output package thành công: {package_name}")
            return package_info
            
        except Exception as e:
            self.log_message(f"❌ Lỗi khi tạo output package: {str(e)}")
            return None
    
    def generate_text_summary(self, output_dir):
        """Tạo text summary file"""
        try:
            # Calculate statistics
            total_entries = len(self.results)
            success_count = len([r for r in self.results if r['status'] == 'success'])
            failed_count = total_entries - success_count
            success_rate = (success_count / total_entries * 100) if total_entries > 0 else 0
            
            # Processing time
            if self.start_time:
                total_time = time.time() - self.start_time
            else:
                total_time = 0
            
            # Error breakdown
            error_breakdown = {}
            for result in self.results:
                if result['status'] == 'failed' and result['error_reason']:
                    error_type = result['error_reason'].split(':')[0]
                    error_breakdown[error_type] = error_breakdown.get(error_type, 0) + 1
            
            # Create summary content
            summary_content = f"""🖼️ IMAGE CRAWLER - BÁO CÁO TÓM TẮT
{'='*60}

📊 THỐNG KÊ TỔNG QUAN:
    • Tổng số entries đã xử lý: {total_entries}
    • Thành công: {success_count} ảnh ({success_rate:.1f}%)
    • Thất bại: {failed_count} ảnh ({100-success_rate:.1f}%)
    • Thời gian xử lý: {total_time:.1f} giây

📁 KẾT QUẢ OUTPUT:
    • Folder ảnh: images/ ({success_count} files)
    • Báo cáo Excel: crawler_report_*.xlsx
    • File tóm tắt: summary.txt (file này)

"""

            if error_breakdown:
                summary_content += "❌ PHÂN TÍCH LỖI:\n"
                for error_type, count in error_breakdown.items():
                    summary_content += f"    • {error_type}: {count} lỗi\n"
                summary_content += "\n"

            if success_count > 0:
                summary_content += "✅ DANH SÁCH ẢNH THÀNH CÔNG:\n"
                for result in self.results:
                    if result['status'] == 'success':
                        size_kb = round(result['file_size'] / 1024, 1) if result['file_size'] else 0
                        summary_content += f"    • {result['filename']} ({size_kb}KB) - {result['product_code']}\n"
                summary_content += "\n"

            if failed_count > 0:
                summary_content += "❌ DANH SÁCH LỖI:\n"
                for result in self.results:
                    if result['status'] == 'failed':
                        summary_content += f"    • {result['product_code']}: {result['error_reason']}\n"

            summary_content += f"\n🕒 Tạo báo cáo: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"

            # Save summary file
            summary_filename = "summary.txt"
            summary_path = os.path.join(output_dir, summary_filename)
            
            with open(summary_path, 'w', encoding='utf-8') as f:
                f.write(summary_content)
            
            self.log_message(f"✅ Đã tạo text summary: {summary_filename}")
            return summary_path
            
        except Exception as e:
            self.log_message(f"❌ Lỗi khi tạo text summary: {str(e)}")
            return None
    
    def stop_crawling(self):
        self.is_crawling = False
        self.log_message("Đang dừng quá trình crawl...")
        
        # Clear queue
        while not self.download_queue.empty():
            try:
                self.download_queue.get_nowait()
                self.download_queue.task_done()
            except queue.Empty:
                break
    
    def crawling_finished(self):
        self.is_crawling = False
        self.start_button.config(state='normal')
        self.stop_button.config(state='disabled')
        self.progress_var.set(100)
        
        # Basic completion message
        basic_message = f"Crawl hoàn thành! Đã xử lý {self.processed_count} entries, thành công {self.success_count}, thất bại {self.failed_count}"
        self.log_message(basic_message)
        
        # Generate comprehensive output package
        if self.output_dir and len(self.results) > 0:
            self.log_message("📊 Bắt đầu tạo báo cáo chi tiết...")
            self.status_label.config(text="Đang tạo báo cáo...")
            
            try:
                # Create output package with reports
                package_info = self.create_output_package(self.output_dir)
                
                if package_info:
                    # Enhanced completion message with package info
                    success_rate = (self.success_count / len(self.results) * 100) if len(self.results) > 0 else 0
                    
                    enhanced_message = f"""🎉 CRAWLER HOÀN THÀNH!

📊 Kết quả tổng kết:
    • Tổng entries: {len(self.results)}
    • Thành công: {self.success_count} ({success_rate:.1f}%)
    • Thất bại: {self.failed_count}

📁 Output package đã tạo:
    • Folder: {package_info['package_name']}
    • Ảnh: {package_info['images_count']} files
    • Báo cáo Excel: ✅
    • Tóm tắt: ✅

📍 Vị trí: {package_info['package_dir']}"""
                    
                    self.log_message("✅ Hoàn thành tạo báo cáo chi tiết!")
                    self.status_label.config(text="Hoàn thành - Có báo cáo!")
                    
                    # Show detailed completion dialog
                    messagebox.showinfo("🎉 Hoàn Thành", enhanced_message)
                    
                    # Optionally open output folder
                    try:
                        import subprocess
                        import platform
                        
                        if platform.system() == "Windows":
                            subprocess.Popen(['explorer', package_info['package_dir']])
                        elif platform.system() == "Darwin":  # macOS
                            subprocess.Popen(['open', package_info['package_dir']])
                        else:  # Linux
                            subprocess.Popen(['xdg-open', package_info['package_dir']])
                    except:
                        pass  # Ignore if can't open folder
                
                else:
                    # Fallback to basic message if package creation failed
                    self.status_label.config(text="Hoàn thành!")
                    messagebox.showinfo("Hoàn thành", basic_message)
            
            except Exception as e:
                self.log_message(f"⚠️ Lỗi khi tạo báo cáo: {str(e)}")
                self.status_label.config(text="Hoàn thành!")
                messagebox.showinfo("Hoàn thành", basic_message)
        
        else:
            # No results to report
            self.status_label.config(text="Hoàn thành!")
            messagebox.showinfo("Hoàn thành", basic_message)
    
    def update_stats(self):
        self.root.after(0, lambda: self.total_links_label.config(text=f"Tổng link: {self.total_links}"))
        self.root.after(0, lambda: self.processed_label.config(text=f"Đã xử lý: {self.processed_count}"))
        self.root.after(0, lambda: self.success_label.config(text=f"Thành công: {self.success_count}"))
        self.root.after(0, lambda: self.failed_label.config(text=f"Thất bại: {self.failed_count}"))
    
    def log_message(self, message):
        timestamp = time.strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}\n"
        
        self.root.after(0, lambda: self.log_text.insert(tk.END, log_entry))
        self.root.after(0, lambda: self.log_text.see(tk.END))

def main():
    root = tk.Tk()
    app = ImageCrawlerApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
